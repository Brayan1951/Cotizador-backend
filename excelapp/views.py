import sys
import os
from django.shortcuts import render
from django.http import HttpResponse
from django.conf import settings
import json
import pandas as pd
from openpyxl import load_workbook
import openpyxl 
from openpyxl.styles import Border, Side
from django.http import JsonResponse
import numpy as np
# exportar excel

def cargar_plantilla():
    if getattr(sys, 'frozen', False):
        script_dir = sys._MEIPASS
    else:

        script_dir = os.path.abspath(".")
    resource_name = './excelapp/formato.xlsx'

    # Obtener la ruta del recurso empaquetado
    resource_path = os.path.join(script_dir, resource_name)
    print(resource_path)
    
    libro=load_workbook(resource_path)
    libro.active
    cotizador=libro["Hoja1"]
    
    return libro,cotizador

def actualizar_plantilla(df_kam,df_cliente,df_coti):
    print(df_kam)
    
    nombre_kam=df_kam["nombre"]
    telefono_kam=df_kam["telefono"]
    correo_kam=df_kam["correo"]
    area_kam=df_kam["area"]
    
    ruc=df_cliente["ruc"]
    nombre=df_cliente["nombre"]
    direccion=df_cliente["direccion"]
    
    libro,temp_plantilla=cargar_plantilla()
    # Establecer un estilo de borde
    borde = Border(left=Side(style='thin'), 
               right=Side(style='thin'), 
               top=Side(style='thin'), 
               bottom=Side(style='thin'))
    
    
    
    suma_sing_igv=df_coti["Total_sin_igv"].sum()
    igv=round(suma_sing_igv*0.18,2)
    monto_total=round(suma_sing_igv*1.18,2)
    
    last_fila=0
    datos_cotizacion=df_coti.values.tolist()
    for index,fila in enumerate(datos_cotizacion,start=26):
        
        for col_num,valor in enumerate(fila,start=2):
            # print({valor})
            # print(type(valor))
            temp_plantilla.cell(row=index,column=col_num).border=borde
            if (col_num<4):
                # temp_plantilla.cell(row=index,column=col_num).border=borde
                temp_plantilla.cell(row=index,column=col_num,value=valor)
            if (col_num==4) :
                temp_plantilla.merge_cells(f'D{index}:G{index}')
                # temp_plantilla.merge_cells(f'D{index}:G{index}').border=borde
                temp_plantilla[f'D{index}']=valor
            if (col_num>4):
                temp_plantilla.cell(row=index,column=col_num+3,value=valor)
                temp_plantilla.cell(row=index,column=col_num+3).border=borde
        last_fila=index+1
        temp_plantilla.insert_rows(index+1)
    temp_plantilla.merge_cells(f'J{last_fila}:K{last_fila}')
    # temp_plantilla.merge_cells(f'J{last_fila}:K{last_fila}').border=borde
    
    for index in range(3):
        temp_plantilla.cell(row=7+index,column=last_fila).border=borde
        
        
        
        
    # agregar datos del cliente
    
    temp_plantilla[f'B{10}']=nombre
    temp_plantilla.merge_cells(f'B{10}:E{10}')
    temp_plantilla[f'B{12}']=ruc
    temp_plantilla.merge_cells(f'B{12}:E{12}')
    temp_plantilla[f'B{14}']=direccion
    temp_plantilla.merge_cells(f'B{14}:E{14}')
    
    # datos kam
    
    temp_plantilla[f'J{19}']=nombre_kam
    temp_plantilla.merge_cells(f'J{19}:L{19}')
    temp_plantilla[f'J{20}']=telefono_kam
    temp_plantilla.merge_cells(f'J{20}:L{20}')
    temp_plantilla[f'J{21}']=correo_kam
    temp_plantilla.merge_cells(f'J{21}:L{21}')
    temp_plantilla[f'J{22}']=area_kam
    temp_plantilla.merge_cells(f'J{22}:L{22}')
    
    
    
    
    
    
    # temp_plantilla[f'J{last_fila}']='Monto sin IGV'
    temp_plantilla[f'L{last_fila+1}']=suma_sing_igv
    
    # temp_plantilla[f'J{last_fila+1}']='IGV'
    temp_plantilla[f'L{last_fila+2}']=igv
    
    # temp_plantilla[f'J{last_fila+2}']='Monto con IGV'
    temp_plantilla[f'L{last_fila+3}']=monto_total
    
    temp_plantilla.merge_cells(f'B{last_fila+5}:L{last_fila+5}')
    temp_plantilla.merge_cells(f'H{last_fila+8}:L{last_fila+11}')
    # cuentas
    temp_plantilla.merge_cells(f'B{last_fila+8}:E{last_fila+8}')
    temp_plantilla.merge_cells(f'B{last_fila+9}:E{last_fila+9}')
    temp_plantilla.merge_cells(f'B{last_fila+10}:E{last_fila+10}')
    temp_plantilla.merge_cells(f'B{last_fila+11}:E{last_fila+11}')
    temp_plantilla.merge_cells(f'B{last_fila+12}:E{last_fila+12}')
    temp_plantilla.merge_cells(f'B{last_fila+13}:E{last_fila+13}')
    temp_plantilla.merge_cells(f'B{last_fila+14}:E{last_fila+14}')
    temp_plantilla.merge_cells(f'B{last_fila+15}:E{last_fila+15}')
    temp_plantilla.merge_cells(f'B{last_fila+16}:E{last_fila+16}')
    
    
    
    
    
    archivo_temporal = 'cotizacion-lista.xlsx'
    
    libro.save('cotizacion-lista.xlsx')

    # return libro
    response = HttpResponse(open(archivo_temporal, 'rb').read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=temp_modificado.xlsx'
    return response

def obtener_data(data):
    data=pd.DataFrame(data)
    # data_modificada
    data['cantidad'] = pd.to_numeric(data['cantidad'], errors='coerce')
    data['precio'] = pd.to_numeric(data['precio'], errors='coerce')
    # data["precio"]=pd.to_numeric(data["precio"])
    data["indice"]=data.index+1
    # data["marca"]=" " 
    # data["precio sin igv"]= round(data["precio"] /(1.18),2)
    data["precio sin igv"] = data["precio"].apply(lambda x: round(x / 1.18, 2) if pd.notna(x) and x != 0 else x)

    # data["Total_sin_igv"]=    round( data["cantidad"]*data["precio"]/(1.18),2)
    # data["Total_sin_igv"] = data.apply(lambda row: round(row["cantidad"] * row["precio"] / 1.18, 2) if pd.notna(row["cantidad"]) and pd.notna(row["precio"]) and row["cantidad"] != 0 and row["precio"] != 0 else np.nan, axis=1)
    data["Total_sin_igv"] = data.apply(
        lambda row: round(row["cantidad"] * row["precio"] / 1.18, 2) 
        if pd.notna(row["cantidad"]) and pd.notna(row["precio"]) and row["cantidad"] != 0 and row["precio"] != 0 
        else np.nan, 
        axis=1
    )
    
    data = data[['indice', 'codigo','descripcion','marca','cantidad','precio sin igv','precio','Total_sin_igv']]
    # print(data)
    return data

# Obtener clientes
def cargar_clientes():
    print("definicion de colums")
    columns_filter=["codigo","ruc","nombre","credito"]
    if getattr(sys, 'frozen', False):
        script_dir = sys._MEIPASS
    else:

        script_dir = os.path.abspath(".")
    # resource_name = './excelapp/cotizador.xlsx'
    resource_name = 'COTIZADOR.xlsx'

        # Obtener la ruta del recurso empaquetado
    
    # resource_path = os.path.join(script_dir,'./excelapp', resource_name)
    resource_path = os.path.join(settings.BASE_DIR,'./excelapp', resource_name)
    print(resource_path)
    print("pasaste por here antes carga")
    clientes=pd.read_excel(resource_path,sheet_name="LC",usecols=columns_filter)
  
        
    print("pasaste por here  carga")
    return clientes
def buscar_clientes(cliente):
    
    temp= str.lower(cliente)
    data_cliente=cargar_clientes()
    print("pasaste por here")
    filter_cliente=data_cliente['nombre'].str.lower().str.startswith(temp)
    print("pasaste por here 2")
    clientes=data_cliente[filter_cliente]
    print("pasaste por here 3")
    lista_clientes=clientes.to_dict(orient='records')
    # lista_clientes=clientes.values.tolist()
    return lista_clientes

def buscar_ruc(ruc):
    data_cliente=cargar_clientes()
    data_cliente['ruc'] = data_cliente['ruc'].astype(str)
    filter_cliente=data_cliente['ruc'].str.startswith(ruc)
    clientes=data_cliente[filter_cliente]
    lista_clientes=clientes.to_dict(orient='records')
    return lista_clientes

# buscar productos

def cargar_productos():
    columns_filter=["codigo_sap","codigo","descripcion","marca"]
    if getattr(sys, 'frozen', False):
            script_dir = sys._MEIPASS
    else:

        script_dir = os.path.abspath(".")
        resource_name = './excelapp/COTIZADOR.xlsx'

        # Obtener la ruta del recurso empaquetado
        resource_path = os.path.join(script_dir, resource_name)
    productos=pd.read_excel(resource_path,"ARTICULOS",usecols=columns_filter)
    return productos

def buscar_productos(codigo):
    temp= str.lower(codigo)
    data_productos=cargar_productos()
    filter_productos=data_productos['codigo'].str.lower().str.startswith(temp,na=False)
    productos=data_productos[filter_productos]
    lista_productos=productos.to_dict(orient='records')
    return lista_productos









# Create your views here.
def hello(request):
    return HttpResponse("Holiwis world")

def generar_excel(request):

        
    if request.method =='GET':
        print("pasaste el get")
        return HttpResponse("estas en el gett")
    else:
        
        
        
        data = json.loads(request.body.decode('utf-8'))
        
        df_kam = data["ejecutivo"]
        df_cliente = data["cliente"]
        df_productos = obtener_data(data["productos"])
        print(df_cliente)
        return actualizar_plantilla(df_kam,df_cliente,df_productos)
    
    
    
def obtener_clientes(request):
    if request.method=='GET':
        print("pasaste el get de obtener clienteS")
        return HttpResponse("estas en el gett")
    else:
        data = json.loads(request.body.decode('utf-8'))
        data_clientes=buscar_clientes(str(data["codigo"]))
        return JsonResponse({'clientes':data_clientes})
        # try:
            
        #     data = json.loads(request.body.decode('utf-8'))
        #     print(data)
        #     data_clientes=buscar_clientes(str(data["codigo"]))
        #     return JsonResponse({'clientes':data_clientes})
        # except Exception:
        #     print(str(Exception))
        #     return HttpResponse(Exception)
            
            
def obtener_cliente_ruc(request,ruc):
    # codigo = request.GET.get('ruc', None)
    data=buscar_ruc(str(ruc))
    print(ruc)
  
    return JsonResponse({'cliente':data})
        
        
def obtener_productos(request):
    if request.method=='GET':
        print("pasaste el get de obtener productos")
        return HttpResponse("estas en el gett productos")
    else:
        data = json.loads(request.body.decode('utf-8'))
        data_productos=buscar_productos(str(data["codigo"]))
        # data_productos=cargar_clientes(data["codigo"])
        # print(data["codigo"])         
        # print(data_productos)
        return JsonResponse({'productos':data_productos})
        
        
    